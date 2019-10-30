from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

def _genResponse(workbook):
    response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
    return response

def genPoItemXlsResponse(responseInput):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = responseInput.query_description.replace('/','') # slash character not allowed in tab title

    worksheet['A1']='PO #'
    worksheet['B1']='Release #'
    worksheet['C1']='Line #'
    worksheet['D1']='P/N'
    worksheet['E1']='Cust. P/N'
    worksheet['F1']='Description'
    worksheet['G1']='Price'
    worksheet['H1']='Qty shipped'
    worksheet['I1']='Total qty'
    worksheet['J1']='Ext. Price'
    worksheet['K1']='Currency'
    worksheet['L1']='Due date [%s thru %s]' % (responseInput.minDate.date(), responseInput.maxDate.date())
    worksheet['M1']='Comment'
    worksheet['N1']='Status'

    row = 2
    for item in responseInput.query_set:
        worksheet.cell(column=1, row=row, value=item.po.num)
        worksheet.cell(column=2, row=row, value=item.po.rel_num)
        worksheet.cell(column=3, row=row, value=item.line_num)
        worksheet.cell(column=4, row=row, value=item.revision.product.pn)
        worksheet.cell(column=5, row=row, value=item.revision.product.cust_pn)
        worksheet.cell(column=6, row=row, value=item.description)
        worksheet.cell(column=7, row=row, value=item.price.price)
        worksheet.cell(column=8, row=row, value=item.shipped_qty)
        worksheet.cell(column=9, row=row, value=item.qty)
        worksheet.cell(column=10, row=row, value=item.extended_price)
        worksheet.cell(column=11, row=row, value=item.price.currency.tla)
        worksheet.cell(column=12, row=row, value=item.due_date)
        worksheet.cell(column=13, row=row, value=item.comment)
        worksheet.cell(column=14, row=row, value=item.status.name)
        row += 1

    return _genResponse(workbook)

def genProductXlsResponse(responseInput):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = responseInput.query_description.replace('/','') # slash character not allowed in tab title

    worksheet['A1']='P/N'
    worksheet['B1']='Customer P/N'
    worksheet['C1']='Description'
    worksheet['D1']='Unit Price'
    worksheet['E1']='Currency'
    worksheet['F1']='MOQ'
    worksheet['G1']='Comments'
    worksheet['H1']='Prod Admin'
    worksheet['I1']='Sales Admin'
    worksheet['J1']='Latest Active Revision'

    row = 2
    for item in responseInput.query_set:
        worksheet.cell(column=1, row=row, value=item.pn)
        worksheet.cell(column=2, row=row, value=item.cust_pn)
        worksheet.cell(column=3, row=row, value=item.description)
        worksheet.cell(column=4, row=row, value=item.price.price)
        worksheet.cell(column=5, row=row, value=item.price.currency.tla)
        worksheet.cell(column=6, row=row, value=item.moq)
        worksheet.cell(column=7, row=row, value=item.comment)
        worksheet.cell(column=8, row=row, value=item.prod_manager.email)
        worksheet.cell(column=9, row=row, value=item.shipping_manager.email)
        worksheet.cell(column=10, row=row, value=item.last_rev)
        row += 1

    return _genResponse(workbook)
